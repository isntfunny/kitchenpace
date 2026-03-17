"""
Convert Swiss Food Composition Database (naehrwertdaten.ch) V 7.0 Excel to JSON.
Adds bundesdeutsche (standard German) names using OpenFoodFacts taxonomy + manual mapping.

Source: https://naehrwertdaten.ch/de/downloads/
License: Free for commercial use with attribution to BLV (Swiss Federal Food Safety and Veterinary Office).

Usage: python3 src/lib/swiss-food-db/convert.py
Prereq: Download OFF taxonomy to /tmp/off_ingredients.json:
  curl -sL -o /tmp/off_ingredients.json https://static.openfoodfacts.org/data/taxonomies/ingredients.json
Output: src/lib/swiss-food-db/data/swiss-foods.json
"""

import json
import os
import re
import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(SCRIPT_DIR, "data", "swiss_food_db.xlsx")
OFF_TAXONOMY_PATH = "/tmp/off_ingredients.json"
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "data", "swiss-foods.json")

# ── Swiss→German word-level replacements ─────────────────────────────────────
# Applied as word boundary replacements in food names.
# Swiss original becomes a synonym.
SWISS_WORD_REPLACEMENTS = {
    "Poulet": "Hähnchen",
    "Broccoli": "Brokkoli",
    "Zucchetti": "Zucchini",
    "Rüebli": "Karotte",
    "Peterli": "Petersilie",
    "Nüsslisalat": "Feldsalat",
    "Federkohl": "Grünkohl",
    "Baumnuss": "Walnuss",
    "Baumnussöl": "Walnussöl",
    "Baumnussbrot": "Walnussbrot",
    "Knöpflimehl": "Spätzlemehl",
    "Gipfeli": "Croissant",
    "Weggli": "Brötchen",
    "Butterweggli": "Butterbrötchen",
    "Semmeli": "Semmel",
    "Mütschli": "Rundbrötchen",
    "Wienerli": "Wiener Würstchen",
    "Guetzli": "Kekse",
    "Schümliguetzli": "Schaumkekse",
    "Ruchbrot": "Ruchbrot",  # keep — known in southern DE, no good equivalent
    "Plätzli": "Schnitzel",
    "Hackplätzli": "Frikadelle",
    "Wähe": "Wähe",  # keep — regional but understood
}

# ── OFF taxonomy rename blocklist ────────────────────────────────────────────
# OFF sometimes gives unhelpful German names. Block these.
OFF_RENAME_BLOCKLIST = {
    "Zucker, weiss",       # OFF says "Saccharose" — nobody searches for that
    "Traubenzucker",       # OFF says "Dextrose" — both are common in DE
    "Schaumwein",          # OFF says "Champagner" — wrong (Champagner is only from Champagne)
    "Bäckerhefe, gepresst",  # OFF says "Bierhefe" — wrong, fresh yeast ≠ brewer's yeast
    "Scholle, roh",        # OFF says "Goldbutt" — "Scholle" is the common German name
    "Seelachs, roh",       # OFF says "Köhler" — "Seelachs" is the standard retail name
    "Apfelwein, 4 vol%",   # OFF says "Apfelmost" — Apfelwein is the DE term
    "Apfelwein, 6.2 vol%",
}

# ── Manual whole-name overrides (Swiss name → German name) ───────────────────
# For cases where neither word replacement nor OFF covers it.
MANUAL_NAME_OVERRIDES = {
    "Vorzugsbutter": "Butter",
    "Sauerrahm": "Crème fraîche",
    "Saurer Halbrahm": "Saure Sahne",
    "Vollrahm, pasteurisiert": "Sahne, pasteurisiert",
    "Vollrahm, UHT": "Sahne, UHT",
    "Doppelrahm, pasteurisiert": "Doppelrahm, pasteurisiert",  # same in DE
    "Halbrahm, pasteurisiert": "Schlagsahne, pasteurisiert",
    "Halbrahm, UHT": "Schlagsahne, UHT",
    "Kaffeerahm": "Kaffeesahne",
    "Rahm (Durchschnitt)": "Sahne (Durchschnitt)",
    "Rahmglace, Aroma": "Sahneeis, Aroma",
    "Rahmglace, Frucht": "Sahneeis, Frucht",
    "Rahmspinat, gekocht": "Rahmspinat, gekocht",  # same in DE
    "Rahmkaramellen": "Sahnekaramellen",
    "Tomatenpüree": "Tomatenmark",
    "Pariserbrot": "Baguette",
    "Fleischkäse": "Leberkäse",
    "Dorsch, roh": "Kabeljau, roh",
    "Egli, roh": "Flussbarsch, roh",
    "Kochsalz mit Jod": "Jodsalz",
    "Kochsalz mit Jod und Fluor": "Speisesalz (jodiert, fluoridiert)",
    "Kochsalz ohne Jod und Fluor": "Speisesalz",
    "Kochspeck": "Speck (durchwachsen)",
    "Rohessspeck": "Bauchspeck, geräuchert",
    "Rüeblitorte": "Karottenkuchen",
    "Konfitüre": "Konfitüre",  # keep — both DE and CH use this
    "Fotzelschnitte ungezuckert": "Armer Ritter, ungezuckert",
}

# ── Nutrient column mapping ──────────────────────────────────────────────────
NUTRIENT_COLUMNS = {
    8: "energyKj",
    11: "energyKcal",
    14: "fat",
    17: "saturatedFat",
    20: "monoUnsaturatedFat",
    23: "polyUnsaturatedFat",
    26: "linoleicAcid",
    29: "alphaLinolenicAcid",
    32: "epa",
    35: "dha",
    38: "cholesterol",
    41: "carbs",
    44: "sugar",
    47: "starch",
    50: "fiber",
    53: "protein",
    56: "salt",
    59: "alcohol",
    62: "water",
    65: "vitaminA_RE",
    68: "vitaminA_RAE",
    71: "retinol",
    74: "betaCaroteneActivity",
    77: "betaCarotene",
    80: "vitaminB1",
    83: "vitaminB2",
    86: "vitaminB6",
    89: "vitaminB12",
    92: "niacin",
    95: "folate",
    98: "pantothenicAcid",
    101: "vitaminC",
    104: "vitaminD",
    107: "vitaminE",
    110: "potassium",
    113: "sodium",
    116: "chloride",
    119: "calcium",
    122: "magnesium",
    125: "phosphorus",
    128: "iron",
    131: "iodine",
    134: "zinc",
    137: "selenium",
}


def parse_numeric(val):
    """Parse a cell value to float, return None for non-numeric values."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s in ("", "k.A.", "-", "Sp.", "Spuren"):
        return None
    try:
        return float(s.replace(",", "."))
    except ValueError:
        return None


def parse_synonyms(val):
    """Parse synonyms field (semicolon-separated) to list."""
    if not val:
        return []
    return [s.strip() for s in str(val).split(";") if s.strip()]


def load_off_taxonomy():
    """Load OpenFoodFacts ingredients taxonomy and build German name lookup."""
    if not os.path.exists(OFF_TAXONOMY_PATH):
        print(f"  OFF taxonomy not found at {OFF_TAXONOMY_PATH}, skipping OFF-based renames")
        return {}

    with open(OFF_TAXONOMY_PATH, encoding="utf-8") as f:
        tax = json.load(f)

    # Build lookup: lowercase German name → preferred German name
    de_lookup = {}
    for key, entry in tax.items():
        names = entry.get("name", {})
        de = names.get("de")
        if de:
            de_lookup[de.lower().strip()] = de.strip()

    print(f"  OFF taxonomy loaded: {len(de_lookup)} German entries")
    return de_lookup


def apply_german_name(swiss_name, synonyms, off_lookup):
    """
    Determine the best bundesdeutsche name for a Swiss food.
    Returns (german_name, updated_synonyms) where swiss original is added as synonym if renamed.
    """
    original = swiss_name
    extra_synonyms = []

    # 1. Manual whole-name override (highest priority)
    if swiss_name in MANUAL_NAME_OVERRIDES:
        german = MANUAL_NAME_OVERRIDES[swiss_name]
        if german != swiss_name:
            extra_synonyms.append(swiss_name)
            return german, list(dict.fromkeys(synonyms + extra_synonyms))

    # 2. Swiss word-level replacements
    replaced = swiss_name
    for ch_word, de_word in SWISS_WORD_REPLACEMENTS.items():
        if ch_word in replaced and ch_word != de_word:
            replaced = replaced.replace(ch_word, de_word)
    if replaced != swiss_name:
        extra_synonyms.append(swiss_name)
        return replaced, list(dict.fromkeys(synonyms + extra_synonyms))

    # 3. OFF taxonomy lookup (check base name before first comma, then full name)
    if swiss_name not in OFF_RENAME_BLOCKLIST:
        base = swiss_name.split(",")[0].strip()
        off_name = off_lookup.get(base.lower())
        if off_name and off_name.lower() != base.lower():
            # Replace the base part, keep qualifiers
            qualifier = swiss_name[len(base):]  # e.g. ", roh"
            german = off_name + qualifier
            extra_synonyms.append(swiss_name)
            return german, list(dict.fromkeys(synonyms + extra_synonyms))

        # Also try synonym match
        for syn in synonyms:
            off_name = off_lookup.get(syn.lower().strip())
            if off_name and off_name.lower() != base.lower():
                qualifier = swiss_name[len(base):]
                german = off_name + qualifier
                extra_synonyms.append(swiss_name)
                return german, list(dict.fromkeys(synonyms + extra_synonyms))

    # 4. No rename needed — name is already standard German
    return swiss_name, synonyms


# ── Cooking-method suffix stripping ────────────────────────────────────────────
# Remove pure cooking/preparation suffixes (heat application, hydration, etc.)
# that create duplicate entries like "Apfel, roh" / "Apfel, gekocht".
# Keeps product-altering states: getrocknet, geräuchert, geschält, geröstet,
# gesalzen, gezuckert, gedörrt, gepresst — these describe different products.
COOK_SUFFIX_RE = re.compile(
    r",\s+"
    r'(?:im Ofen |"medium" |im Wasser )?'
    r"(?:"
    r"roh"
    r"|gekocht"
    r"|gebraten"
    r"|gedämpft"
    r"|gedünstet"
    r"|geschmort"
    r"|gebacken"
    r"|gefroren"
    r"|gefriergetrocknet"
    r"|frittiert"
    r"|zubereitet"
    r"|abgetropft"
    r"|aufgewärmt"
    r"|tiefgekühlt"
    r"|ungebacken"
    r"|ungebraten"
    r"|paniert und \w+"
    r")"
    r"(?:\s*\(.*\))?"        # optional parenthetical e.g. (ohne Zugabe von Salz)
    r"(?:\s+(?:in|im)\s+.*)?"  # optional trailing e.g. "in HOLL-Rapsöl"
    r"$"
)


def strip_cooking_suffix(name):
    """Strip cooking-method suffix from a food name. Returns cleaned name."""
    return COOK_SUFFIX_RE.sub("", name)


def convert():
    print(f"Reading {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws = wb["Generische Lebensmittel"]

    # Verify header row
    header_row = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    assert header_row[0] == "ID"
    assert header_row[3] == "Name"
    spot_checks = {
        8: "Energie, Kilojoule (kJ)",
        14: "Fett, total (g)",
        41: "Kohlenhydrate, verfügbar (g)",
        53: "Protein (g)",
    }
    for col_idx, expected in spot_checks.items():
        actual = header_row[col_idx]
        assert actual == expected, f"Column {col_idx}: expected '{expected}', got '{actual}'"

    print("Loading OFF taxonomy for German name lookup...")
    off_lookup = load_off_taxonomy()

    foods = []
    rename_count = 0
    for row in ws.iter_rows(min_row=4, values_only=True):
        swiss_id = row[0]
        swiss_name = row[3]
        if not swiss_name or not swiss_id:
            continue

        swiss_name = str(swiss_name).strip()
        synonyms = parse_synonyms(row[4])
        category = (row[5] or "").strip()
        density = parse_numeric(row[6])

        nutrients = {}
        for col_idx, key in NUTRIENT_COLUMNS.items():
            val = parse_numeric(row[col_idx])
            if val is not None:
                nutrients[key] = val

        # Apply German name mapping
        german_name, updated_synonyms = apply_german_name(swiss_name, synonyms, off_lookup)
        if german_name != swiss_name:
            rename_count += 1

        # Strip cooking-method suffixes (e.g. ", roh", ", gekocht")
        clean_name = strip_cooking_suffix(german_name)
        if clean_name != german_name:
            # Keep original as synonym so it's still searchable
            if german_name not in updated_synonyms:
                updated_synonyms = updated_synonyms + [german_name]

        foods.append({
            "id": int(swiss_id),
            "name": clean_name,
            "swissName": swiss_name if german_name != swiss_name else None,
            "synonyms": updated_synonyms,
            "category": category,
            "density": density,
            "nutrients": nutrients,
        })

    # Deduplicate by name (after cooking-suffix stripping, many names collide)
    pre_dedup = len(foods)
    seen = {}
    deduped = []
    for food in foods:
        key = food["name"].lower()
        if key in seen:
            # Merge synonyms into first occurrence
            existing = seen[key]
            merged = list(dict.fromkeys(existing["synonyms"] + food["synonyms"]))
            existing["synonyms"] = merged
            # Merge categories (semicolon-separated, deduplicated)
            existing_cats = set(c.strip() for c in existing["category"].split(";") if c.strip())
            new_cats = set(c.strip() for c in food["category"].split(";") if c.strip())
            existing["category"] = "; ".join(sorted(existing_cats | new_cats))
        else:
            seen[key] = food
            deduped.append(food)
    foods = deduped
    stripped_count = pre_dedup - len(foods)

    # Sort by name for deterministic output
    foods.sort(key=lambda f: f["name"].lower())

    # Remove null swissName fields for cleaner JSON
    for food in foods:
        if food["swissName"] is None:
            del food["swissName"]

    print(f"Parsed {pre_dedup} foods, {stripped_count} duplicates removed after cooking-suffix stripping")
    print(f"Final: {len(foods)} unique foods ({rename_count} renamed to Bundesdeutsch)")

    # Write JSON
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(foods, f, ensure_ascii=False, indent=2)
    print(f"Written to {OUTPUT_PATH}")

    # Show all renames
    renamed = [f for f in foods if "swissName" in f]
    print(f"\n--- Renames ({len(renamed)}) ---")
    for f in sorted(renamed, key=lambda x: x["swissName"]):
        print(f"  {f['swissName']:55s} → {f['name']}")

    # Stats
    print(f"\n--- Nutrient coverage ({len(foods)} foods) ---")
    for key in ["energyKcal", "fat", "saturatedFat", "carbs", "sugar", "protein", "salt", "fiber"]:
        count = sum(1 for food in foods if key in food["nutrients"])
        print(f"  {key:30s}: {count:4d} ({count/len(foods)*100:.1f}%)")

    big7 = ["energyKcal", "fat", "saturatedFat", "carbs", "sugar", "protein", "salt"]
    big7_complete = sum(1 for f in foods if all(k in f["nutrients"] for k in big7))
    print(f"\n  Big 7 complete: {big7_complete} / {len(foods)} ({big7_complete/len(foods)*100:.1f}%)")

    wb.close()


if __name__ == "__main__":
    convert()
